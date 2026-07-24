from datetime import UTC, datetime, timedelta
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock

from openpyxl import load_workbook

from app.constants import SUBJECTS
from app.data.default_students import DEFAULT_STUDENTS
from app.models import Branch, Competition, Student, Subject
from app.services.excel_service import export_top3_results


def _student(
    student_id: int,
    full_name: str,
    branch_name: str,
    subject_name: str,
    category: str,
):
    first_name, last_name = full_name.split(maxsplit=1)
    return SimpleNamespace(
        id=student_id,
        first_name=first_name,
        last_name=last_name,
        full_name=full_name,
        category=category,
        branch=SimpleNamespace(name=branch_name),
        subject=SimpleNamespace(name=subject_name),
        subject_id=1 if subject_name == "IT" else 2,
    )


def test_it_dasturlash_is_merged_into_it_roster():
    assert "IT-dasturlash" not in SUBJECTS
    assert all(subject != "IT-dasturlash" for _name, _branch, subject, _grade in DEFAULT_STUDENTS)

    it_students = [
        (name, branch, grade)
        for name, branch, subject, grade in DEFAULT_STUDENTS
        if subject == "IT"
    ]
    assert ("ABDUJAMALOV ABDULLOH", "Kasblar", 6) in it_students
    assert ("TOJIBOYEVA DILNURA", "Kasblar", 7) in it_students
    assert any(grade <= 6 for _name, _branch, grade in it_students)
    assert any(grade >= 7 for _name, _branch, grade in it_students)


async def test_top3_excel_has_subject_and_grade_winners():
    students_and_votes = [
        (_student(1, "Ali Valiyev", "Kasblar", "IT", "1-6"), 10),
        (_student(2, "Vali Aliyev", "Chinoz", "IT", "1-6"), 8),
        (_student(3, "Sami Karimov", "Gulbahor", "IT", "1-6"), 6),
        (_student(4, "Tomi Salimov", "Olmazor", "IT", "1-6"), 2),
        (_student(5, "Lola Ergasheva", "Kasblar", "IT", "7-11"), 5),
        (_student(6, "Nodir Omonov", "Chinoz", "IT", "7-11"), 3),
        (_student(7, "Zafar Sobirov", "Gulbahor", "IT", "7-11"), 1),
        (_student(8, "Olim Qodirov", "Olmazor", "IT", "7-11"), 0),
    ]
    result = MagicMock()
    result.all.return_value = [
        SimpleNamespace(Student=student, vote_count=votes)
        for student, votes in students_and_votes
    ]
    session = AsyncMock()
    session.execute.return_value = result

    now = datetime.now(UTC)
    competition = SimpleNamespace(
        id=1,
        name="Test tanlov",
        state="finished",
        starts_at=now - timedelta(days=7),
        ends_at=now,
    )
    content = await export_top3_results(session, competition)
    workbook = load_workbook(BytesIO(content), data_only=False)

    assert workbook.sheetnames == [
        "1-2-3 orinlar",
        "1-6-sinflar",
        "7-11-sinflar",
        "IT",
    ]

    ws = workbook["1-2-3 orinlar"]
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    assert len(rows) == 6
    assert rows[0][1:6] == (1, "Ali Valiyev", "Kasblar", "IT", "1–6-sinflar")
    assert rows[1][1] == 2
    assert rows[2][1] == 3
    assert rows[3][1:6] == (1, "Lola Ergasheva", "Kasblar", "IT", "7–11-sinflar")
    assert rows[0][8] == "=IF(H5=0,0,G5/H5)"

    # Har bir fan alohida varaqda bo‘ladi va 0 ovozli qatnashchi ham qoladi.
    it_ws = workbook["IT"]
    it_rows = list(it_ws.iter_rows(min_row=5, values_only=True))
    assert len(it_rows) == 8
    assert it_rows[0][2:6] == (
        "Ali Valiyev",
        "Kasblar",
        "IT",
        "1–6-sinflar",
    )
    assert it_rows[-1][1] == "-"
    assert it_rows[-1][2] == "Olim Qodirov"
    assert it_rows[-1][6:9] == (0, 9, "=IF(H12=0,0,G12/H12)")
