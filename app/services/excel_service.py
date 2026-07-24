from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.constants import CATEGORY_LABELS, SUBJECTS
from app.models import Branch, Competition, Student, Subject, Vote
from app.services.admin_service import category_for_grade
from app.utils.time import competition_status, format_local

IMPORT_HEADERS = ["Ism", "Familiya", "Filial", "Fan", "Sinf"]
REPORT_HEADERS = [
    "№",
    "O‘rin",
    "Ism",
    "Familiya",
    "Filial",
    "Fan",
    "Sinf",
    "Sinf toifasi",
    "Ovozlar soni",
    "Foiz",
    "Tanlov nomi",
]

TOP3_HEADERS = [
    "№",
    "O‘rin",
    "F.I.SH",
    "Filial",
    "Fan",
    "Sinf",
    "Ovozlar soni",
    "Guruh jami ovozi",
    "Foiz",
    "Tanlov nomi",
]


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    widths = [6, 10, 18, 18, 20, 20, 12, 18, 16, 12, 32]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width


def build_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Oquvchilar"
    ws.append(IMPORT_HEADERS)
    ws.append(["Ali", "Valiyev", "Niyozbosh", "Ingliz tili", 4])
    ws.append(["Madina", "Karimova", "Gulbahor", "Matematika", 7])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="2F75B5")
        cell.font = Font(color="FFFFFF", bold=True)
    for col, width in zip("ABCDE", [18, 18, 22, 22, 10], strict=True):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


async def import_students(
    session: AsyncSession, content: bytes
) -> tuple[int, bytes | None]:
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError("Excel faylni o‘qib bo‘lmadi") from exc
    ws = wb.active
    headers = [
        str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]
    ]
    if headers[:5] != IMPORT_HEADERS:
        raise ValueError(
            "Excel ustunlari: Ism, Familiya, Filial, Fan, Sinf bo‘lishi kerak"
        )

    branches = {
        b.name.casefold(): b for b in (await session.scalars(select(Branch))).all()
    }
    subjects = {
        s.name.casefold(): s for s in (await session.scalars(select(Subject))).all()
    }
    existing_rows = (
        await session.execute(
            select(
                func.lower(Student.first_name),
                func.lower(Student.last_name),
                Student.branch_id,
                Student.subject_id,
                Student.grade,
            )
        )
    ).all()
    existing = {(a, b, c, d, e) for a, b, c, d, e in existing_rows}

    errors: list[list[object]] = []
    added = 0
    for row_number, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        first_name, last_name, branch_name, subject_name, grade_raw = (
            list(row[:5]) + [None] * 5
        )[:5]
        row_values = [first_name, last_name, branch_name, subject_name, grade_raw]
        try:
            if not all(
                (
                    first_name,
                    last_name,
                    branch_name,
                    subject_name,
                    grade_raw is not None,
                )
            ):
                raise ValueError("Majburiy katak bo‘sh")
            branch = branches.get(str(branch_name).strip().casefold())
            subject_key = str(subject_name).strip().casefold()
            if subject_key in {"it-dasturlash", "it dasturlash", "it_dasturlash"}:
                subject_key = "it"
            subject = subjects.get(subject_key)
            if branch is None:
                raise ValueError("Filial topilmadi")
            if subject is None:
                raise ValueError("Fan topilmadi")
            grade = int(grade_raw)
            category = category_for_grade(grade)
            key = (
                str(first_name).strip().lower(),
                str(last_name).strip().lower(),
                branch.id,
                subject.id,
                grade,
            )
            if key in existing:
                raise ValueError("Takroriy o‘quvchi")
            session.add(
                Student(
                    first_name=str(first_name).strip(),
                    last_name=str(last_name).strip(),
                    branch_id=branch.id,
                    subject_id=subject.id,
                    grade=grade,
                    category=category,
                    is_active=True,
                )
            )
            existing.add(key)
            added += 1
        except Exception as exc:
            errors.append([row_number, *row_values, str(exc)])

    await session.commit()
    if not errors:
        return added, None

    error_wb = Workbook()
    error_ws = error_wb.active
    error_ws.title = "Xatolar"
    error_ws.append(["Qator", *IMPORT_HEADERS, "Xato sababi"])
    for error in errors:
        error_ws.append(error)
    for cell in error_ws[1]:
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.font = Font(color="FFFFFF", bold=True)
    for col, width in zip("ABCDEFG", [10, 18, 18, 22, 22, 10, 32], strict=True):
        error_ws.column_dimensions[col].width = width
    output = BytesIO()
    error_wb.save(output)
    return added, output.getvalue()


async def export_results(session: AsyncSession, competition: Competition) -> bytes:
    students = list(
        (
            await session.scalars(
                select(Student)
                .options(selectinload(Student.branch), selectinload(Student.subject))
                .order_by(
                    Student.branch_id,
                    Student.subject_id,
                    Student.category,
                    Student.last_name,
                )
            )
        ).all()
    )
    vote_counts = dict(
        (
            await session.execute(
                select(Vote.student_id, func.count(Vote.id))
                .where(Vote.competition_id == competition.id)
                .group_by(Vote.student_id)
            )
        ).all()
    )

    group_totals: dict[tuple[int, int, str], int] = {}
    for student in students:
        key = (student.branch_id, student.subject_id, student.category)
        group_totals[key] = group_totals.get(key, 0) + int(
            vote_counts.get(student.id, 0)
        )

    grouped: dict[tuple[int, int, str], list[Student]] = {}
    for student in students:
        grouped.setdefault(
            (student.branch_id, student.subject_id, student.category), []
        ).append(student)

    rank_map: dict[int, int] = {}
    for group_students in grouped.values():
        ordered = sorted(
            group_students,
            key=lambda item: (
                -int(vote_counts.get(item.id, 0)),
                item.last_name,
                item.first_name,
            ),
        )
        previous_votes = None
        rank = 0
        for index, student in enumerate(ordered, start=1):
            votes = int(vote_counts.get(student.id, 0))
            if previous_votes is None or votes < previous_votes:
                rank = index
            rank_map[student.id] = rank
            previous_votes = votes

    def rows_for(items: list[Student]) -> list[list[object]]:
        rows: list[list[object]] = []
        ordered = sorted(
            items,
            key=lambda s: (
                s.branch.name,
                s.subject.name,
                s.category,
                rank_map.get(s.id, 9999),
                s.last_name,
            ),
        )
        for number, student in enumerate(ordered, start=1):
            votes = int(vote_counts.get(student.id, 0))
            total = group_totals[
                (student.branch_id, student.subject_id, student.category)
            ]
            percent = round(votes / total * 100, 1) if total else 0.0
            rows.append(
                [
                    number,
                    rank_map.get(student.id, 0),
                    student.first_name,
                    student.last_name,
                    student.branch.name,
                    student.subject.name,
                    student.grade,
                    CATEGORY_LABELS[student.category],
                    votes,
                    percent,
                    competition.name,
                ]
            )
        return rows

    wb = Workbook()
    wb.remove(wb.active)

    sheets: list[tuple[str, list[Student]]] = [
        ("Umumiy natijalar", students),
        ("Filiallar bo‘yicha", students),
        (
            "Fanlar bo‘yicha",
            sorted(students, key=lambda s: (s.subject.name, s.branch.name)),
        ),
        ("1-6-sinflar", [s for s in students if s.category == "1-6"]),
        ("7-11-sinflar", [s for s in students if s.category == "7-11"]),
    ]
    branch_names = []
    for student in students:
        if student.branch.name not in branch_names:
            branch_names.append(student.branch.name)
    for branch_name in branch_names:
        sheets.append(
            (branch_name[:31], [s for s in students if s.branch.name == branch_name])
        )

    used_names: set[str] = set()
    for title, items in sheets:
        base = title[:31]
        name = base
        suffix = 1
        while name in used_names:
            suffix += 1
            name = f"{base[:27]}-{suffix}"
        used_names.add(name)
        ws = wb.create_sheet(name)
        ws.append(REPORT_HEADERS)
        for row in rows_for(items):
            ws.append(row)
        _style_sheet(ws)
        for cell in ws["J"][1:]:
            cell.number_format = "0.0%"
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value / 100

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


async def export_top3_results(session: AsyncSession, competition: Competition) -> bytes:
    """1–2–3-o‘rinlar va har bir fan bo‘yicha barcha o‘quvchilar Exceli."""
    stmt = (
        select(Student, func.count(Vote.id).label("vote_count"))
        .outerjoin(
            Vote,
            (Vote.student_id == Student.id)
            & (Vote.competition_id == competition.id),
        )
        .options(selectinload(Student.branch), selectinload(Student.subject))
        .where(Student.is_active.is_(True))
        .group_by(Student.id)
    )
    raw_rows = (await session.execute(stmt)).all()

    grouped: dict[tuple[int, str], list[tuple[Student, int]]] = {}
    for row in raw_rows:
        student = row.Student
        grouped.setdefault((student.subject_id, student.category), []).append(
            (student, int(row.vote_count))
        )

    subject_order = {name: index for index, name in enumerate(SUBJECTS)}
    category_order = {"1-6": 0, "7-11": 1}
    all_rows: list[dict[str, object]] = []
    winner_rows: list[dict[str, object]] = []

    ordered_groups = sorted(
        grouped.values(),
        key=lambda items: (
            subject_order.get(items[0][0].subject.name, len(subject_order)),
            category_order.get(items[0][0].category, 99),
            items[0][0].subject.name.casefold(),
        ),
    )
    for items in ordered_groups:
        total_votes = sum(votes for _student, votes in items)
        ordered = sorted(
            items,
            key=lambda item: (
                -item[1],
                item[0].branch.name.casefold(),
                item[0].last_name.casefold(),
                item[0].first_name.casefold(),
            ),
        )

        previous_votes: int | None = None
        current_rank = 0
        for index, (student, votes) in enumerate(ordered, start=1):
            if previous_votes is None or votes < previous_votes:
                current_rank = index
            previous_votes = votes
            # Ovoz olmagan qatnashchiga sovrinli o‘rin yozilmaydi.
            place: int | str = current_rank if votes > 0 else "-"
            row_data: dict[str, object] = {
                "place": place,
                "student": student,
                "votes": votes,
                "total": total_votes,
            }
            all_rows.append(row_data)
            if votes > 0 and current_rank <= 3:
                winner_rows.append(row_data)

    wb = Workbook()
    wb.remove(wb.active)

    status = competition_status(
        competition.state, competition.starts_at, competition.ends_at
    )
    status_text = "Yakuniy natijalar" if status == "finished" else "Joriy natijalar"

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    place_fills = {
        1: PatternFill("solid", fgColor="FFF2CC"),
        2: PatternFill("solid", fgColor="E7E6E6"),
        3: PatternFill("solid", fgColor="FCE4D6"),
    }

    def unique_sheet_name(title: str, used_names: set[str]) -> str:
        invalid = '[]:*?/\\'
        cleaned = "".join("-" if char in invalid else char for char in title).strip()
        base = (cleaned or "Natijalar")[:31]
        name = base
        suffix = 1
        while name.casefold() in used_names:
            suffix += 1
            tail = f"-{suffix}"
            name = f"{base[:31 - len(tail)]}{tail}"
        used_names.add(name.casefold())
        return name

    used_names: set[str] = set()

    def create_result_sheet(
        title: str,
        heading: str,
        rows: list[dict[str, object]],
        empty_text: str,
    ) -> None:
        ws = wb.create_sheet(unique_sheet_name(title, used_names))
        ws.merge_cells("A1:J1")
        ws["A1"] = heading
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="17365D")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:J2")
        ws["A2"] = (
            f"{competition.name} | {status_text} | "
            f"Tugash: {format_local(competition.ends_at)}"
        )
        ws["A2"].font = Font(italic=True, color="404040")
        ws["A2"].alignment = Alignment(horizontal="center")

        header_row = 4
        for col, value in enumerate(TOP3_HEADERS, start=1):
            cell = ws.cell(row=header_row, column=col, value=value)
            cell.fill = PatternFill("solid", fgColor="2F75B5")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        if not rows:
            ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=10)
            ws.cell(5, 1, empty_text)
            ws.cell(5, 1).alignment = Alignment(horizontal="center")
            ws.cell(5, 1).font = Font(italic=True, color="7F6000")
            ws.cell(5, 1).border = border
        else:
            for number, item in enumerate(rows, start=1):
                student = item["student"]
                votes = int(item["votes"])
                total = int(item["total"])
                row_index = header_row + number
                values = [
                    number,
                    item["place"],
                    student.full_name,
                    student.branch.name,
                    student.subject.name,
                    CATEGORY_LABELS[student.category],
                    votes,
                    total,
                    None,
                    competition.name,
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_index, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(
                        horizontal="center" if col in {1, 2, 6, 7, 8, 9} else "left",
                        vertical="center",
                        wrap_text=True,
                    )
                ws.cell(
                    row_index,
                    9,
                    f'=IF(H{row_index}=0,0,G{row_index}/H{row_index})',
                )
                ws.cell(row_index, 9).number_format = "0.0%"
                place = item["place"]
                if isinstance(place, int) and place in place_fills:
                    for cell in ws[row_index]:
                        cell.fill = place_fills[place]

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:J{max(4, 4 + len(rows))}"
        widths = [6, 10, 30, 20, 22, 18, 16, 18, 12, 38]
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + index)].width = width

    # Umumiy sovrinli o‘rinlar saqlanadi.
    create_result_sheet(
        "1-2-3 orinlar",
        "AL-AZIZ ACADEMY — FAN VA SINF BO‘YICHA 1–2–3-O‘RINLAR",
        winner_rows,
        "Hozircha ovoz olgan o‘quvchilar mavjud emas.",
    )
    create_result_sheet(
        "1-6-sinflar",
        "AL-AZIZ ACADEMY — 1–6-SINFLAR 1–2–3-O‘RINLARI",
        [item for item in winner_rows if item["student"].category == "1-6"],
        "1–6-sinflarda hozircha ovoz olgan o‘quvchilar mavjud emas.",
    )
    create_result_sheet(
        "7-11-sinflar",
        "AL-AZIZ ACADEMY — 7–11-SINFLAR 1–2–3-O‘RINLARI",
        [item for item in winner_rows if item["student"].category == "7-11"],
        "7–11-sinflarda hozircha ovoz olgan o‘quvchilar mavjud emas.",
    )

    # Har bir fan alohida varaq: barcha faol qatnashchilar va foizlari.
    subject_names = sorted(
        {item["student"].subject.name for item in all_rows},
        key=lambda name: (subject_order.get(name, len(subject_order)), name.casefold()),
    )
    for subject_name in subject_names:
        subject_rows = [
            item for item in all_rows if item["student"].subject.name == subject_name
        ]
        subject_rows.sort(
            key=lambda item: (
                category_order.get(item["student"].category, 99),
                -int(item["votes"]),
                item["student"].branch.name.casefold(),
                item["student"].last_name.casefold(),
                item["student"].first_name.casefold(),
            )
        )
        create_result_sheet(
            subject_name,
            f"AL-AZIZ ACADEMY — {subject_name.upper()} NATIJALARI",
            subject_rows,
            f"{subject_name} fanida faol o‘quvchilar mavjud emas.",
        )

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

